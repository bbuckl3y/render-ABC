"""
Program to extract Booklets from ABC library files

Sets to extract come from a MS Excel file
Each row has a comma separated list of tune X IDs in column A

Bob Buckley 3/12/2025
"""

import sys
import ABClib
import openpyxl

def main():
    # read an ABC file 
    # file names should come from arguments 
    # fn = "PBBtunes20251103.abc"
    # wb = "abcextract-test.xlsx"
    # fd = "XXX.abc"

    if len(sys.argv)!=4:
        print("usage: prog SrcABC xlsx DstABC")
        exit(1)
    
    pn, fn, wb, fd = sys.argv
    assert fn.endswith(".abc")
    assert wb.endswith(".xlsx")
    assert fd.endswith(".abc")

    ss=ABClib.Songsets(fn)
    td = ABClib.abcdict(ss.abcs()) # read tune dictionary
    hdr = ss.hdr
    ss = None
    
    dwsa = openpyxl.load_workbook(wb).active
    # active workbook in Excel file - may need more here
    cols = list(dwsa.iter_cols(1, dwsa.max_column))
    setlist = [ str(c.value).split('-') for c in cols[0] ]

    if setlist[0][0].startswith('set'): # A1 starts with 'set' if it's a header line
        setlist.pop(0)
    # print("setlist=", setlist)

    # separators - %%newpage before and after a set, %%sep between single tunes
    seps = [ "\n\n%%newpage\n" if any(len(x)>1 for x in xs) else "\n\n%%sep\n" for xs in zip(setlist, setlist[1:]) ]+[""]
    
    with open(fd, "tw") as dst:
        for l in hdr:# write ABC file header
            dst.write(l)
            dst.write('\n')
        for vs, s in zip(setlist, seps):
            # print('-'.join(vs))
            for tasc in vs:
                dst.write("\nX: "+tasc+"\n")
                if tasc in td:
                    for l in td[tasc]:
                        dst.write(l)
                        dst.write('\n')
                else:
                    print('tune', tasc, 'not in ABC file.')
            if s:
                dst.write(s)
        dst.write('\n\n')
    return

if __name__=="__main__":
    main()